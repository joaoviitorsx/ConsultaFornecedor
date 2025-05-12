import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from tkinter import filedialog, messagebox
import os

def salvar_resultados(lista_dados, caminho_padrao='resultado.xlsx'):
    print("[DEBUG] Iniciando processo de salvamento...")

    if not lista_dados:
        print("[DEBUG] Nenhum dado recebido para salvar.")
        messagebox.showwarning("Aviso", "Nenhum dado foi gerado para salvar.")
        return

    print(f"[DEBUG] {len(lista_dados)} registros a serem salvos.")

    resposta_salvar = messagebox.askyesno("Salvar Planilha", "Deseja salvar a planilha gerada?")
    print(f"[DEBUG] Resposta do usuário para salvar: {resposta_salvar}")

    if not resposta_salvar:
        print("[DEBUG] Usuário optou por não salvar.")
        return

    caminho_arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Planilha Excel", "*.xlsx")],
        initialfile=caminho_padrao,
        title="Salvar como"
    )

    print(f"[DEBUG] Caminho selecionado: {caminho_arquivo}")
    if not caminho_arquivo:
        print("[DEBUG] Caminho de salvamento não fornecido.")
        return

    try:
        colunas_desejadas = ['Código', 'CNPJ', 'Razão Social', 'Simples Nacional', 'Decreto']
        df = pd.DataFrame(lista_dados)

        df = df[[col for col in colunas_desejadas if col in df.columns]]

        for col in df.columns:
            df[col] = df[col].astype(str)

        df.to_excel(caminho_arquivo, index=False)

        wb = load_workbook(caminho_arquivo)
        ws = wb.active

        fonte_arial = Font(name='Arial', size=11)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = fonte_arial
                cell.number_format = "@"

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(caminho_arquivo)
        print("[DEBUG] Planilha salva com sucesso.")

        resposta_abrir = messagebox.askyesno("Abrir Planilha", "Deseja abrir a planilha agora?")
        print(f"[DEBUG] Resposta do usuário para abrir: {resposta_abrir}")

        if resposta_abrir and os.path.exists(caminho_arquivo):
            os.startfile(caminho_arquivo)

        return caminho_arquivo

    except Exception as e:
        print(f"[DEBUG][ERRO] Ocorreu uma exceção: {str(e)}")
        messagebox.showerror("Erro", f"Ocorreu um erro ao salvar a planilha:\n{str(e)}")
        return None
